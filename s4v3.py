#import openpyxl
from s4v2 import *
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.utils import get_column_letter

def save_spreadsheet(filename, data_sample):
    wb = Workbook()
    ws = wb.active

    row_index = 1
    for rows in data_sample:        #rows
        col_index = 1
        for field in rows:          #columns
            col_letter = get_column_letter(col_index)
            ws.cell('{}{}'.format(col_letter, row_index)).value = field   #like A1 in spreadsheet
            col_index +=1
        row_index +=1

    wb.save(filename)

kiton_ties = filter_col_by_by_string(data_from_csv, "brandName", "Kiton")   #new data sample
save_spreadsheet("_data/s4-kiton.xlsx", kiton_ties)

